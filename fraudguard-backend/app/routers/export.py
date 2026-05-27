"""
app/routers/export.py
─────────────────────────────────────────────────────────────────────────────
Export endpoints:
  GET /api/v1/export/run/{run_id}   → download Excel report for one run
─────────────────────────────────────────────────────────────────────────────
Generates a 3-sheet workbook:
  Sheet 1 — Summary    (run metadata + AI summary)
  Sheet 2 — Procedures (one row per procedure result)
  Sheet 3 — Flagged Items (all flags with full detail)
─────────────────────────────────────────────────────────────────────────────
"""

from __future__ import annotations

import io
import uuid
from datetime import datetime

import asyncpg
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.database import get_db
from app.models.schemas import UserOut
from app.routers.auth import get_current_user

router = APIRouter(prefix="/export", tags=["export"])


# ── Style helpers ─────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _thin_border() -> Border:
    s = Side(style="thin", color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)


HEADER_FILL = _fill("1E3A5F")
HEADER_FONT = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
LABEL_FONT  = Font(name="Calibri", bold=True, size=10, color="374151")
VALUE_FONT  = Font(name="Calibri", size=10)
TITLE_FONT  = Font(name="Calibri", bold=True, size=14, color="1E3A5F")

RISK_FILL = {
    "High":   _fill("FEE2E2"),
    "Medium": _fill("FEF3C7"),
    "Low":    _fill("DCFCE7"),
}
DEFAULT_FILL = _fill("F9FAFB")


def _header_row(ws, headers: list[str], col_widths: list[int]):
    """Write a styled header row and set column widths."""
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(1, col, h)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.border    = _thin_border()
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = w


# ── Main export route ─────────────────────────────────────────────────────────

@router.get("/run/{run_id}")
async def export_run(
    run_id: str,
    current_user: UserOut = Depends(get_current_user),
    db: asyncpg.Connection = Depends(get_db),
):
    # ── Fetch data ────────────────────────────────────────────────────────────
    run = await db.fetchrow(
        """SELECT id, file_name, file_size, row_count,
                  started_at, completed_at, status, ai_summary
           FROM analysis_runs WHERE id = $1""",
        uuid.UUID(run_id),
    )
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")

    proc_rows = await db.fetch(
        """SELECT procedure_name, status, risk_level, flag_count
           FROM procedure_results
           WHERE run_id = $1
           ORDER BY created_at""",
        uuid.UUID(run_id),
    )

    flag_rows = await db.fetch(
        """SELECT row_id, invoice_no, vendor_id, amount, date,
                  procedure_name, reason, risk_level, document_type,
                  field, flagged_value, detection, status, auditor_action
           FROM flagged_items
           WHERE run_id = $1
           ORDER BY
               CASE risk_level
                   WHEN 'High'   THEN 1
                   WHEN 'Medium' THEN 2
                   WHEN 'Low'    THEN 3
                   ELSE 4
               END,
               created_at""",
        uuid.UUID(run_id),
    )

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.column_dimensions["A"].width = 26
    ws1.column_dimensions["B"].width = 52

    ws1["A1"] = "FraudGuard — Analysis Report"
    ws1["A1"].font = TITLE_FONT
    ws1.merge_cells("A1:B1")
    ws1.row_dimensions[1].height = 28

    summary_rows = [
        ("Generated",      datetime.utcnow().strftime("%d %b %Y  %H:%M UTC")),
        ("File",           run["file_name"]),
        ("File Size",      run["file_size"] or "—"),
        ("Row Count",      run["row_count"] or "—"),
        ("Status",         run["status"].capitalize()),
        ("Started",        run["started_at"].strftime("%d %b %Y  %H:%M") if run["started_at"] else "—"),
        ("Completed",      run["completed_at"].strftime("%d %b %Y  %H:%M") if run["completed_at"] else "—"),
        ("Procedures Run", len(proc_rows)),
        ("Total Flags",    len(flag_rows)),
        ("High-Risk Flags",sum(1 for f in flag_rows if f["risk_level"] == "High")),
    ]

    for i, (label, value) in enumerate(summary_rows, 2):
        ws1.cell(i, 1, label).font = LABEL_FONT
        ws1.cell(i, 2, value).font = VALUE_FONT

    # AI summary (can be long)
    ai_row = len(summary_rows) + 3
    ws1.cell(ai_row, 1, "AI Summary").font = LABEL_FONT
    ai_cell = ws1.cell(ai_row, 2, run["ai_summary"] or "—")
    ai_cell.font      = VALUE_FONT
    ai_cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws1.row_dimensions[ai_row].height = 72

    # ── Sheet 2: Procedures ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("Procedures")
    _header_row(ws2, ["Procedure", "Status", "Risk Level", "Flags Found"], [38, 14, 14, 14])

    for i, p in enumerate(proc_rows, 2):
        risk   = p["risk_level"] or ""
        fill   = RISK_FILL.get(risk, DEFAULT_FILL)
        values = [
            p["procedure_name"],
            p["status"].capitalize(),
            risk or "—",
            p["flag_count"],
        ]
        for col, v in enumerate(values, 1):
            cell           = ws2.cell(i, col, v)
            cell.font      = VALUE_FONT
            cell.fill      = fill
            cell.border    = _thin_border()
            cell.alignment = Alignment(
                horizontal="left" if col == 1 else "center"
            )

    # ── Sheet 3: Flagged Items ────────────────────────────────────────────────
    ws3 = wb.create_sheet("Flagged Items")

    flag_headers = [
        "Row ID", "Invoice No", "Vendor ID", "Amount", "Date",
        "Procedure", "Reason", "Risk Level", "Document Type",
        "Field", "Flagged Value", "Detection", "Status", "Auditor Action",
    ]
    flag_widths = [10, 16, 12, 14, 12, 30, 44, 12, 16, 18, 18, 14, 14, 22]
    _header_row(ws3, flag_headers, flag_widths)

    for i, f in enumerate(flag_rows, 2):
        risk = f["risk_level"] or ""
        fill = RISK_FILL.get(risk, DEFAULT_FILL)
        values = [
            f["row_id"],
            f["invoice_no"],
            f["vendor_id"],
            f["amount"],
            f["date"],
            f["procedure_name"],
            f["reason"],
            risk or "—",
            f["document_type"],
            f["field"],
            f["flagged_value"],
            f["detection"],
            f["status"],
            f["auditor_action"],
        ]
        for col, v in enumerate(values, 1):
            cell           = ws3.cell(i, col, v or "—")
            cell.font      = VALUE_FONT
            cell.fill      = fill
            cell.border    = _thin_border()
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws3.row_dimensions[i].height = 26

    # Freeze header + auto-filter
    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = f"A1:{get_column_letter(len(flag_headers))}1"

    # ── Stream to client ──────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = run["file_name"].replace(" ", "_").replace("/", "-")
    date_str  = datetime.utcnow().strftime("%Y%m%d")
    filename  = f"fraudguard_{safe_name}_{date_str}.xlsx"

    return StreamingResponse(
        buf,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"'
        },
    )